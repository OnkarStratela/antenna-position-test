"""Antenna-position test logger.

Drives the existing `rfid_reader` C binary one throw at a time and records
every throw to a single master Excel workbook (`results.xlsx`) with two
sheets:

- ``Throws``    : one row per throw — setup, throw number, timing, number of
  unique tags, EPC list, antennas hit, plus an embedded thumbnail of the
  antenna-position photo for that setup.
- ``TagReads``  : one row per detected tag — setup, throw, EPC, antenna, ts.

The C reader dedupes EPCs for its entire process lifetime, so this script
spawns a *fresh* `rfid_reader` subprocess for every throw and kills it when
the operator ends the throw. That way the same container can be thrown again
in a later throw and will still be reported.

The WS2812 LED strip + GPIO13 PWM LED behaviour from ``rfid_led.py`` is
preserved here: solid WHITE while idle, a green blink for every new unique
tag detected during a throw, OFF on exit. The LED layer is optional — if
``rpi_ws281x`` / ``gpiozero`` aren't available (or the script isn't run
with sudo), the logger still records throws but skips the LEDs.

Operator workflow (no SSH-and-restart between throws)::

    $ sudo python3 antenna_test_logger.py

    Available antenna setups:
      1. antennas_opposite
      2. antennas_vertical
      3. antennas_horizontal
    Select starting setup [1/2/3]: 1

    [Setup: antennas_opposite]  ENTER = start throw  |  's' = switch setup  |  'q' = quit
    > <ENTER>
    [antennas_opposite | Throw #1] starting reader... (press ENTER once you've thrown the containers to end)
        [RFID] TAG DETECTED: ABCDEF1234 [Source_0] [2026-05-26 16:10:00]
    > <ENTER>
    [antennas_opposite | Throw #1] DONE — 1 unique tag(s), 3.2 s
        logged to results.xlsx
"""

from __future__ import annotations

import atexit
import datetime as dt
import os
import queue
import re
import signal
import subprocess
import sys
import threading
import time
from dataclasses import dataclass, field
from pathlib import Path
from typing import List, Optional

from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter
from PIL import Image as PILImage

try:
    from rpi_ws281x import PixelStrip, Color  # WS2812 strip on GPIO12
except Exception:  # pragma: no cover — host without rpi_ws281x still runs
    PixelStrip = None  # type: ignore[assignment]
    Color = None  # type: ignore[assignment]

try:
    from gpiozero import PWMLED  # GPIO13 PWM LED held on during the session
except Exception:  # pragma: no cover
    PWMLED = None  # type: ignore[assignment]


SCRIPT_DIR    = Path(__file__).resolve().parent
RFID_BINARY   = SCRIPT_DIR / "rfid_reader"
IMAGES_DIR    = SCRIPT_DIR / "images"
RESULTS_XLSX  = SCRIPT_DIR / "results.xlsx"
THUMB_DIR     = SCRIPT_DIR / ".thumbs"

SETUPS: List[str] = [
    "antennas_opposite",
    "antennas_vertical",
    "antennas_horizontal",
]

# rfid_reader.c emits each new EPC on a line like:
#   \x1b[0;32m[RFID] TAG DETECTED: ABCDEF12 \x1b[0m [Source_0] [2026-05-26 16:00:00]
# The ANSI colour codes are stripped before matching.
ANSI_RE = re.compile(r"\x1b\[[0-9;]*m")
TAG_LINE_RE = re.compile(
    r"\[RFID\] TAG DETECTED:\s*(?P<epc>[0-9A-Fa-f]+)\s*"
    r"\[(?P<antenna>[^\]]+)\]\s*\[(?P<ts>[^\]]+)\]"
)
# The C reader prints this once it has finished setup and is actively
# polling both antennas. We wait for it before declaring a throw "live".
READER_READY_RE = re.compile(r"\[RFID\] Scanning on")

THROWS_HEADERS = [
    "session_id",
    "setup",
    "throw_num",
    "start_time",
    "end_time",
    "duration_s",
    "n_unique_tags",
    "epcs",
    "antennas_hit",
    "setup_photo",
    "thrown_count",   # operator-reported ground truth for this throw
    "hit_rate_pct",   # n_unique_tags / thrown_count * 100, when known
]

TAGREADS_HEADERS = [
    "session_id",
    "setup",
    "throw_num",
    "epc",
    "antenna",
    "timestamp",
]

# Column widths (in Excel character units) for readability.
THROWS_WIDTHS    = [16, 22, 10, 22, 22, 11, 8, 60, 22, 20, 13, 13]
TAGREADS_WIDTHS  = [16, 22, 10, 28, 12, 22]

# Embedded-thumbnail height in pixels. Excel row height is in points; one
# row is sized to comfortably hold an 80-px-tall image.
THUMB_HEIGHT_PX  = 80
ROW_HEIGHT_PT    = 64


# ─────────────────────── LED feedback (optional) ────────────────────────
#
# Mirrors rfid_led.py: WS2812 strip on GPIO12 + a held-on PWM LED on GPIO13.
# Idle = solid WHITE; each new unique tag = brief WHITE off-pulse → GREEN
# for GREEN_HOLD_SECONDS → back to WHITE. Multiple new tags in quick
# succession produce N visible blinks. On exit, the strip is turned OFF
# and GPIO13 is forced LOW at the SoC pin-mux level (so the PWM LED can't
# latch HIGH after cleanup).
#
# All of this is best-effort: if rpi_ws281x or gpiozero aren't available,
# the logger keeps running and just skips the LEDs.

LED_COUNT       = 19
LED_PIN         = 12
LED_FREQ_HZ     = 600_000
LED_DMA         = 10
LED_BRIGHTNESS  = 225
LED_INVERT      = False
LED_CHANNEL     = 0

PWM_LED_PIN         = 13
PWM_LED_BRIGHTNESS  = 0.5

GREEN_HEX = "#00FF00"
WHITE_HEX = "#FFFFFF"
OFF_HEX   = "#000000"

GREEN_HOLD_SECONDS  = 1.0
WHITE_FLASH_SECONDS = 0.10


def _force_gpio_low(pin: int) -> None:
    """Hardware-level final LOW for the GPIO13 PWM LED.

    Last-resort cleanup: gpiozero / lgpio sometimes leave the SoC pin-mux
    register configured as OUTPUT-HIGH when they release the GPIO chardev
    claim on Pi OS Bookworm — which makes the LED snap to maximum
    brightness as the program exits. ``pinctrl`` / ``raspi-gpio`` writes
    the SoC pin-mux register directly, and that state persists after our
    process exits.
    """
    for cmd in (
        ["pinctrl", "set", str(pin), "op", "dl"],
        ["raspi-gpio", "set", str(pin), "op", "dl"],
    ):
        try:
            subprocess.run(
                cmd,
                check=True,
                stdout=subprocess.DEVNULL,
                stderr=subprocess.DEVNULL,
                timeout=2,
            )
            return
        except Exception:
            continue


def _hex_to_color(hex_code: str):
    r = int(hex_code[1:3], 16)
    g = int(hex_code[3:5], 16)
    b = int(hex_code[5:7], 16)
    return Color(r, g, b)


class LEDFeedback:
    """WS2812 + GPIO13 PWM LED driver, behaving exactly like rfid_led.py
    but spanning the entire test session (across all throws). Safe to
    construct on any host: degrades to a no-op if the libraries can't
    initialise (e.g. no rpi_ws281x, missing /dev/mem permissions, …)."""

    def __init__(self) -> None:
        self.enabled = False
        self._strip = None
        self._pwm_led = None
        self._queue: "queue.Queue[float]" = queue.Queue()
        self._stop = threading.Event()
        self._worker: Optional[threading.Thread] = None
        self._GREEN = self._WHITE = self._OFF = None

        if PixelStrip is None or Color is None:
            print("[LED] rpi_ws281x not available; running without LED feedback.")
            return

        try:
            self._strip = PixelStrip(
                LED_COUNT, LED_PIN, LED_FREQ_HZ,
                LED_DMA, LED_INVERT, LED_BRIGHTNESS, LED_CHANNEL,
            )
            self._strip.begin()
            self._GREEN = _hex_to_color(GREEN_HEX)
            self._WHITE = _hex_to_color(WHITE_HEX)
            self._OFF   = _hex_to_color(OFF_HEX)
            self._fill(self._WHITE)
            self.enabled = True
        except Exception as exc:
            print(f"[LED] WARN: could not init WS2812 strip ({exc}); "
                  f"continuing without LEDs. (Run with sudo for WS2812 access.)")
            self._strip = None
            return

        # GPIO13 PWM LED: register the SoC-level cleanup BEFORE constructing
        # PWMLED so it runs AFTER gpiozero's own atexit handler (LIFO order).
        atexit.register(_force_gpio_low, PWM_LED_PIN)
        if PWMLED is not None:
            try:
                self._pwm_led = PWMLED(PWM_LED_PIN)
                self._pwm_led.value = PWM_LED_BRIGHTNESS
                print(
                    f"[LED] GPIO{PWM_LED_PIN} PWM LED ON at "
                    f"{int(PWM_LED_BRIGHTNESS * 100)}% duty."
                )
            except Exception as exc:
                print(f"[LED] WARN: could not init GPIO{PWM_LED_PIN} PWM LED: {exc}")
                self._pwm_led = None

        self._worker = threading.Thread(target=self._run_worker, daemon=True)
        self._worker.start()

    def notify_tag(self) -> None:
        """Called for every NEW unique tag detected. Triggers one visible
        green blink (or restarts the current one if we're still green)."""
        if self.enabled:
            self._queue.put(time.time())

    def shutdown(self) -> None:
        """Idempotent. Turns the strip OFF, releases the PWM LED, and
        forces GPIO13 LOW at the SoC level."""
        if self._stop.is_set():
            return
        self._stop.set()
        if self._worker is not None:
            self._worker.join(timeout=2)
        if self._strip is not None and self._OFF is not None:
            try:
                self._fill(self._OFF)
            except Exception:
                pass
        if self._pwm_led is not None:
            try:
                self._pwm_led.value = 0.0
                time.sleep(0.05)
            except Exception:
                pass
            try:
                self._pwm_led.close()
            except Exception:
                pass
        _force_gpio_low(PWM_LED_PIN)

    def _fill(self, color) -> None:
        assert self._strip is not None
        for i in range(self._strip.numPixels()):
            self._strip.setPixelColor(i, color)
        self._strip.show()

    def _run_worker(self) -> None:
        assert self._strip is not None
        self._fill(self._WHITE)
        while not self._stop.is_set():
            try:
                self._queue.get(timeout=0.1)
            except queue.Empty:
                continue
            # New tag → produce one green blink.
            self._fill(self._WHITE)
            time.sleep(WHITE_FLASH_SECONDS)
            self._fill(self._GREEN)
            green_until = time.time() + GREEN_HOLD_SECONDS
            while not self._stop.is_set():
                remaining = green_until - time.time()
                if remaining <= 0:
                    break
                try:
                    self._queue.get(timeout=remaining)
                except queue.Empty:
                    break
                # Another tag during the green window → restart blink.
                self._fill(self._WHITE)
                time.sleep(WHITE_FLASH_SECONDS)
                self._fill(self._GREEN)
                green_until = time.time() + GREEN_HOLD_SECONDS
            self._fill(self._WHITE)


# ─────────────────────────── data types ────────────────────────────


@dataclass
class TagRead:
    epc: str
    antenna: str
    ts: str


@dataclass
class ThrowResult:
    setup: str
    throw_num: int
    start_time: dt.datetime
    end_time: dt.datetime
    reads: List[TagRead] = field(default_factory=list)
    # Filled in by the operator after the throw ends; None if they skipped.
    thrown_count: Optional[int] = None

    @property
    def duration_s(self) -> float:
        return (self.end_time - self.start_time).total_seconds()

    @property
    def unique_epcs(self) -> List[str]:
        seen: List[str] = []
        for r in self.reads:
            if r.epc not in seen:
                seen.append(r.epc)
        return seen

    @property
    def antennas_hit(self) -> List[str]:
        return sorted({r.antenna for r in self.reads})

    @property
    def hit_rate_pct(self) -> Optional[float]:
        """n_unique_tags / thrown_count, in percent, rounded to 1 dp.
        Returns None when the operator didn't report a count, or reported 0."""
        if self.thrown_count is None or self.thrown_count <= 0:
            return None
        return round(len(self.unique_epcs) / self.thrown_count * 100, 1)


# ─────────────────────────── workbook I/O ───────────────────────────


def ensure_workbook() -> None:
    """Create `results.xlsx` with the two sheets if it doesn't already exist,
    or upgrade an older file in place if new columns have been added to
    THROWS_HEADERS / TAGREADS_HEADERS since the file was created. Header
    cells that are already populated are left alone (so manual renames
    survive); only empty header cells get filled in."""
    if not RESULTS_XLSX.exists():
        wb = Workbook()
        throws = wb.active
        throws.title = "Throws"
        throws.append(THROWS_HEADERS)
        for col_idx, w in enumerate(THROWS_WIDTHS, start=1):
            throws.column_dimensions[get_column_letter(col_idx)].width = w

        tagreads = wb.create_sheet("TagReads")
        tagreads.append(TAGREADS_HEADERS)
        for col_idx, w in enumerate(TAGREADS_WIDTHS, start=1):
            tagreads.column_dimensions[get_column_letter(col_idx)].width = w

        wb.save(RESULTS_XLSX)
        return

    wb = load_workbook(RESULTS_XLSX)
    changed = False

    def upgrade(sheet_name: str, headers: List[str], widths: List[int]) -> bool:
        nonlocal changed
        if sheet_name not in wb.sheetnames:
            return False
        ws = wb[sheet_name]
        local_changed = False
        for i, h in enumerate(headers, start=1):
            existing = ws.cell(row=1, column=i).value
            if existing is None or existing == "":
                ws.cell(row=1, column=i, value=h)
                local_changed = True
        for col_idx, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(col_idx)].width = w
        if local_changed:
            changed = True
        return local_changed

    upgrade("Throws", THROWS_HEADERS, THROWS_WIDTHS)
    upgrade("TagReads", TAGREADS_HEADERS, TAGREADS_WIDTHS)

    if changed:
        wb.save(RESULTS_XLSX)


def load_existing_throw_counts() -> dict:
    """Read the Throws sheet and return ``{setup: total_throws_logged}`` so
    the per-setup throw counter resumes from where the last session left
    off (instead of resetting to 0 every run)."""
    counts = {s: 0 for s in SETUPS}
    if not RESULTS_XLSX.exists():
        return counts
    try:
        wb = load_workbook(RESULTS_XLSX, read_only=True)
    except Exception:
        return counts
    if "Throws" not in wb.sheetnames:
        wb.close()
        return counts
    ws = wb["Throws"]
    rows = ws.iter_rows(min_row=2, values_only=True)
    for row in rows:
        if not row:
            continue
        # column layout: session_id, setup, throw_num, ...
        setup = row[1] if len(row) > 1 else None
        if setup in counts:
            counts[setup] += 1
    wb.close()
    return counts


def format_counts(counts: dict) -> str:
    return " | ".join(f"{s}: {counts.get(s, 0)}" for s in SETUPS)


def thumb_for(setup: str) -> Optional[Path]:
    """Return a cached thumbnail path for `setup`, regenerating if the
    source photo has been updated. Returns None if no source image exists."""
    src = IMAGES_DIR / f"{setup}.png"
    if not src.exists():
        return None
    THUMB_DIR.mkdir(exist_ok=True)
    dst = THUMB_DIR / f"{setup}_thumb.png"
    if dst.exists() and dst.stat().st_mtime >= src.stat().st_mtime:
        return dst
    img = PILImage.open(src)
    img.thumbnail((10_000, THUMB_HEIGHT_PX))
    img.save(dst, "PNG")
    return dst


def append_throw(session_id: str, throw: ThrowResult) -> None:
    """Append `throw` to results.xlsx: one summary row in Throws (with the
    setup-photo thumbnail embedded) and one row per tag read in TagReads."""
    wb = load_workbook(RESULTS_XLSX)
    throws = wb["Throws"]
    tagreads = wb["TagReads"]

    photo_col_letter = get_column_letter(len(THROWS_HEADERS))
    next_row = throws.max_row + 1
    throws.append([
        session_id,
        throw.setup,
        throw.throw_num,
        throw.start_time.strftime("%Y-%m-%d %H:%M:%S"),
        throw.end_time.strftime("%Y-%m-%d %H:%M:%S"),
        round(throw.duration_s, 2),
        len(throw.unique_epcs),
        ", ".join(throw.unique_epcs),
        ", ".join(throw.antennas_hit),
        "",
        throw.thrown_count if throw.thrown_count is not None else "",
        throw.hit_rate_pct if throw.hit_rate_pct is not None else "",
    ])
    throws.row_dimensions[next_row].height = ROW_HEIGHT_PT

    thumb = thumb_for(throw.setup)
    if thumb is not None:
        img = XLImage(str(thumb))
        img.anchor = f"{photo_col_letter}{next_row}"
        throws.add_image(img)

    for r in throw.reads:
        tagreads.append([
            session_id,
            throw.setup,
            throw.throw_num,
            r.epc,
            r.antenna,
            r.ts,
        ])

    wb.save(RESULTS_XLSX)


# ────────────────────── rfid_reader subprocess ──────────────────────


def run_one_throw(setup: str, throw_num: int, led: LEDFeedback) -> ThrowResult:
    """Spawn `rfid_reader`, stream its output, collect every reported tag
    detection until the operator presses ENTER, then kill the subprocess
    and return the collected ThrowResult. Each new tag also drives a green
    blink on the LED strip."""
    print(
        f"\n[{setup} | Throw #{throw_num}] starting reader... "
        f"(press ENTER once you've thrown the containers to end)"
    )

    proc = subprocess.Popen(
        [str(RFID_BINARY)],
        stdout=subprocess.PIPE,
        stderr=subprocess.STDOUT,
        bufsize=1,
        universal_newlines=True,
    )

    reads: List[TagRead] = []
    reader_ready = threading.Event()
    start_time: List[dt.datetime] = []  # mutable slot set by reader thread

    def reader_thread() -> None:
        assert proc.stdout is not None
        for line in proc.stdout:
            sys.stdout.write("    " + line)
            sys.stdout.flush()
            clean = ANSI_RE.sub("", line)
            if not reader_ready.is_set() and READER_READY_RE.search(clean):
                start_time.append(dt.datetime.now())
                reader_ready.set()
                print(
                    f"[{setup} | Throw #{throw_num}] LIVE — drop containers, "
                    f"press ENTER to end."
                )
                sys.stdout.flush()
            m = TAG_LINE_RE.search(clean)
            if m:
                reads.append(TagRead(
                    epc=m.group("epc"),
                    antenna=m.group("antenna"),
                    ts=m.group("ts"),
                ))
                led.notify_tag()

    t = threading.Thread(target=reader_thread, daemon=True)
    t.start()

    try:
        input()  # block until operator presses ENTER (or EOF/Ctrl-D)
    except (KeyboardInterrupt, EOFError):
        pass

    end_time = dt.datetime.now()

    # Send SIGINT so rfid_reader's signal handler runs CAENRFID_Disconnect()
    # cleanly. Escalate if it doesn't comply.
    try:
        proc.send_signal(signal.SIGINT)
        proc.wait(timeout=3)
    except Exception:
        try:
            proc.terminate()
            proc.wait(timeout=2)
        except Exception:
            proc.kill()
    t.join(timeout=2)

    # If reader_ready never fired (e.g. reader couldn't connect), fall back
    # to "now" so we still log the failed throw.
    actual_start = start_time[0] if start_time else end_time

    return ThrowResult(
        setup=setup,
        throw_num=throw_num,
        start_time=actual_start,
        end_time=end_time,
        reads=reads,
    )


# ───────────────────────────── menu loop ─────────────────────────────


def pick_setup() -> int:
    """Prompt the operator for a setup; returns an index into SETUPS."""
    while True:
        print("\nAvailable antenna setups:")
        for i, s in enumerate(SETUPS, 1):
            print(f"  {i}. {s}")
        choice = input("Select setup [1/2/3]: ").strip()
        if choice in {"1", "2", "3"}:
            return int(choice) - 1
        print("  Invalid choice, try again.")


def prompt_thrown_count() -> Optional[int]:
    """Ask the operator how many containers/tags they just threw. Returns
    the integer they entered, or None if they pressed ENTER to skip or hit
    EOF/Ctrl-C. Re-prompts on invalid input."""
    while True:
        try:
            raw = input(
                "    How many containers did you throw? (ENTER to skip): "
            ).strip()
        except (KeyboardInterrupt, EOFError):
            return None
        if not raw:
            return None
        try:
            n = int(raw)
        except ValueError:
            print(f"    '{raw}' isn't a whole number. Try again, or ENTER to skip.")
            continue
        if n < 0:
            print("    Must be 0 or positive. Try again, or ENTER to skip.")
            continue
        return n


def main() -> int:
    if not RFID_BINARY.is_file() or not os.access(RFID_BINARY, os.X_OK):
        print(f"ERROR: '{RFID_BINARY}' not found or not executable.")
        print("Build it first:  ./compile.sh")
        return 1

    ensure_workbook()
    session_id = dt.datetime.now().strftime("%Y%m%d-%H%M%S")
    print(f"=== Antenna-position test session {session_id} ===")
    print(f"Results: {RESULTS_XLSX}")

    led = LEDFeedback()

    # Make sure LEDs are turned off on any abnormal exit (uncaught
    # exception, SIGTERM, etc). Normal exit goes through the `finally`
    # block below.
    def _signal_shutdown(*_args) -> None:
        led.shutdown()
        sys.exit(0)
    signal.signal(signal.SIGTERM, _signal_shutdown)

    try:
        # Per-setup throw counters resume from previously-logged throws so
        # the live tally is cumulative across runs (not reset every start).
        throw_counters = load_existing_throw_counts()
        print(f"[Live count] {format_counts(throw_counters)}")

        setup_idx = pick_setup()

        while True:
            setup = SETUPS[setup_idx]
            prompt = (
                f"\n[Setup: {setup}]  ENTER = start throw  |  "
                f"'s' = switch setup  |  'q' = quit\n> "
            )
            try:
                choice = input(prompt).strip().lower()
            except (KeyboardInterrupt, EOFError):
                print("\nBye.")
                return 0

            if choice == "q":
                print("Bye.")
                return 0
            if choice == "s":
                setup_idx = pick_setup()
                continue

            throw_counters[setup] += 1
            result = run_one_throw(setup, throw_counters[setup], led)
            print(
                f"[{setup} | Throw #{result.throw_num}] DONE — "
                f"{len(result.unique_epcs)} unique tag(s), "
                f"{result.duration_s:.1f} s"
            )
            result.thrown_count = prompt_thrown_count()
            if result.hit_rate_pct is not None:
                print(
                    f"    {len(result.unique_epcs)}/{result.thrown_count} "
                    f"detected = {result.hit_rate_pct:.1f}% hit rate"
                )
            try:
                append_throw(session_id, result)
                print(f"    logged to {RESULTS_XLSX.name}")
            except Exception as exc:
                print(f"    ERROR writing to {RESULTS_XLSX.name}: {exc}")
                print("    (throw was NOT saved — fix the issue and retry)")
            print(f"[Live count] {format_counts(throw_counters)}")
    finally:
        led.shutdown()


if __name__ == "__main__":
    sys.exit(main())
